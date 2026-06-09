#!/usr/bin/env python3
"""
Generate the functional test plan for Hermanas as a multi-sheet Excel workbook.
Output: docs/test-plan/hermanas-test-plan.xlsx

Layout:
  - "Cover"       : how to use the plan + role legend + environment setup
  - "Roles"       : exact permissions of Guest / USER / ADMIN
  - "Test cases"  : one row per test, one column per role with the expected outcome
                    + columns for actual result and pass/fail status
  - "Smoke"       : short post-deploy smoke test (10 quick checks across roles)
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


# ─── Styling helpers ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E78")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def style_data_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = WRAP_TOP
        cell.border = BORDER


def style_section_row(ws: Worksheet, row: int, ncols: int, text: str) -> None:
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = BORDER
    ws.row_dimensions[row].height = 22


def set_col_widths(ws: Worksheet, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── Sheet: Cover ─────────────────────────────────────────────────────────────

def build_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("Cover")

    title = ws.cell(row=1, column=1, value="Hermanas — Plan de test fonctionnel")
    title.font = Font(name="Calibri", size=18, bold=True, color="1F4E78")
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value="Objectif : faire le tour de l'application avec les 3 rôles (Guest, USER, ADMIN) "
                                    "et vérifier que les contrôles attendus s'affichent et fonctionnent.").alignment = WRAP_TOP
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 32

    # Environment preparation
    style_section_row(ws, 4, 4, "Préparation de l'environnement")
    prep_rows = [
        ("1", "Démarrer le backend Spring Boot (port 8080) — `mvn spring-boot:run` ou JAR déployé"),
        ("2", "Avoir un fichier users.properties avec au moins 1 compte USER et 1 compte ADMIN (bcrypt). "
              "Générer un hash : `java -jar hermanas.jar --hash <password>`."),
        ("3", "Lancer le SPA Angular si non bundlé : `cd frontend && npm start` (sinon ouvrir directement la page servie par Spring)."),
        ("4", "Avoir Chrome / Firefox avec DevTools ouvert (onglet Network) pour vérifier les codes HTTP."),
        ("5", "Recommandé : tester en navigation privée pour repartir d'une session vierge entre chaque rôle."),
    ]
    row = 5
    for n, txt in prep_rows:
        ws.cell(row=row, column=1, value=n).alignment = CENTER
        ws.cell(row=row, column=2, value=txt).alignment = WRAP_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    # Conventions
    style_section_row(ws, row + 1, 4, "Conventions")
    row += 2
    conventions = [
        ("✅", "Affichage / action attendu et fonctionnel"),
        ("👁️", "Affichage en lecture seule"),
        ("⛔", "Élément masqué ou action refusée (HTTP 401/403)"),
        ("🔁", "Login modal proposée à l'utilisateur"),
        ("—", "Sans objet pour ce rôle"),
    ]
    for sym, meaning in conventions:
        ws.cell(row=row, column=1, value=sym).alignment = CENTER
        ws.cell(row=row, column=2, value=meaning).alignment = WRAP_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    # Pass/Fail values
    style_section_row(ws, row + 1, 4, "Statuts de test")
    row += 2
    statuses = [
        ("OK", "Comportement conforme à l'attendu"),
        ("KO", "Comportement différent — détailler dans la colonne 'Commentaire'"),
        ("N/A", "Non testable / hors périmètre pour cette session"),
        ("Blocked", "Test bloqué par un préalable (à reprendre plus tard)"),
    ]
    for sym, meaning in statuses:
        ws.cell(row=row, column=1, value=sym).alignment = CENTER
        ws.cell(row=row, column=2, value=meaning).alignment = WRAP_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    # How to run
    style_section_row(ws, row + 1, 4, "Mode opératoire recommandé")
    row += 2
    steps = [
        "1. Faire l'intégralité du parcours en Guest (sans se connecter).",
        "2. Se connecter avec un compte USER → re-parcourir les écrans listés.",
        "3. Se déconnecter, se reconnecter avec un compte ADMIN → re-parcourir.",
        "4. Pour chaque test, remplir 'Résultat <rôle>' avec OK / KO / N/A et ajouter un commentaire si KO.",
        "5. À la fin, faire passer la feuille 'Smoke' en post-déploiement pour valider que rien n'a régressé.",
    ]
    for s in steps:
        ws.cell(row=row, column=1, value=s).alignment = WRAP_TOP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    set_col_widths(ws, [6, 35, 35, 35])


# ─── Sheet: Roles ─────────────────────────────────────────────────────────────

def build_roles(wb: Workbook) -> None:
    ws = wb.create_sheet("Roles")
    headers = ["Aspect", "Guest (anonyme)", "USER", "ADMIN"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers))

    rows = [
        ("Authentification requise", "Non — accès direct", "Oui — login + mot de passe", "Oui — login + mot de passe"),
        ("Menu : Dashboard", "✅ visible", "✅ visible", "✅ visible"),
        ("Menu : Webcam (/camera)", "⛔ masqué (guard AuthGuard)", "✅ visible", "✅ visible"),
        ("Menu : Music / Weather / Residents / Electronics / Logs / System",
         "✅ visible", "✅ visible", "✅ visible"),
        ("Menu : Users (/notification)", "⛔ masqué", "✅ visible", "✅ visible"),
        ("Menu : Energy saving (/energy)", "⛔ masqué", "⛔ masqué (AdminGuard)", "✅ visible"),
        ("Photos de la webcam (historique)", "⛔ 401 sur /api/v1/camera/photos/**", "✅ accès", "✅ accès"),
        ("Snapshot en direct + analyse IA",
         "👁️ snapshot OK, analyse IA refusée", "✅ snapshot + analyse IA", "✅ snapshot + analyse IA"),
        ("Modifier la configuration (/api/v1/config/**)",
         "⛔ 401", "✅ autorisé", "✅ autorisé"),
        ("Ouvrir / fermer la porte, switches lampe/ventilo/musique",
         "⛔ 401 → login modal", "✅ autorisé", "✅ autorisé"),
        ("Réglage servomoteur, calibration",
         "⛔", "✅ autorisé (USER+ADMIN)", "✅ autorisé"),
        ("Reboot / shutdown machine",
         "⛔ 403", "⛔ 403", "✅ autorisé"),
        ("Page Logs : voir journal métier (événements porte/lampe/etc.)",
         "✅ visible (events business public)", "✅", "✅"),
        ("Page Logs : voir journal auth + fichiers de log",
         "⛔ panneaux masqués", "⛔ panneaux masqués", "✅ visible"),
        ("Page Electronique : table GPIO + image board",
         "✅ visible", "✅ visible", "✅ visible"),
        ("Page Electronique : état direct boutons + relais + servo calibration",
         "⛔ (états masqués / 401)", "⛔ panneau servo & boutons masqués", "✅ visible"),
        ("Page Système : panneau À propos + état système (version)",
         "✅ visible", "✅ visible", "✅ visible"),
        ("Page Système : disk usage + menu actions (reboot/shutdown/refresh)",
         "⛔ masqué", "⛔ masqué", "✅ visible"),
        ("Page Notification / Users : gestion utilisateurs",
         "⛔ AuthGuard redirige", "👁️ profil seul (pas d'admin)", "✅ admin: liste/valider/supprimer"),
    ]

    for i, r in enumerate(rows, start=2):
        for j, v in enumerate(r, start=1):
            ws.cell(row=i, column=j, value=v)
        style_data_row(ws, i, len(headers))
        ws.row_dimensions[i].height = 30

    set_col_widths(ws, [38, 38, 38, 38])
    ws.freeze_panes = "A2"


# ─── Sheet: Test cases ────────────────────────────────────────────────────────

# Layout: section, id, page/url, scenario, pre-req, steps, expected guest, expected user, expected admin,
# result guest, result user, result admin, comment

TEST_HEADERS = [
    "Section", "ID", "Page / URL", "Scénario", "Pré-requis", "Étapes",
    "Attendu Guest", "Attendu USER", "Attendu ADMIN",
    "Résultat Guest", "Résultat USER", "Résultat ADMIN",
    "Commentaire",
]


def _t(*vals):
    """Compact helper to build a test row tuple."""
    return list(vals) + ["", "", "", ""]  # 4 trailing empty cells for results + comment


TEST_DATA: list[tuple[str, list[list]]] = [
    ("Authentification", [
        _t("AUTH-01", "/auth/login", "Connexion avec un compte USER valide",
           "Backend démarré, compte USER existant",
           "1) Cliquer sur l'icône utilisateur en haut à droite ▸ 'Se connecter' "
           "2) Saisir login + mot de passe 3) Valider",
           "🔁 Login proposée puis dashboard rechargé",
           "✅ Login réussi — redirection sur la page courante, nom du compte affiché dans le footer du menu",
           "✅ Idem USER + accès aux entrées admin du menu"),
        _t("AUTH-02", "/auth/login", "Connexion avec mauvais mot de passe",
           "Compte existant",
           "1) Ouvrir le login modal 2) Saisir un mot de passe incorrect 3) Valider",
           "Message d'erreur 'Identifiant ou mot de passe invalide.'",
           "Idem",
           "Idem"),
        _t("AUTH-03", "/auth/register", "Inscription d'un nouveau compte",
           "—",
           "1) Cliquer 'Créer un compte' depuis le login modal 2) Remplir le formulaire 3) Soumettre",
           "✅ Compte créé en NOT_VALIDATED_YET, login refusé tant que pas validé par admin",
           "Idem",
           "Idem"),
        _t("AUTH-04", "—", "Déconnexion",
           "Être connecté",
           "1) Cliquer sur le menu utilisateur 2) 'Se déconnecter'",
           "—",
           "✅ Footer du menu repasse à 'Guest', éléments admin/USER masqués sans F5",
           "Idem"),
        _t("AUTH-05", "—", "Persistance de session (remember-me)",
           "Compte USER ou ADMIN",
           "1) Se connecter avec 'Se souvenir de moi' coché 2) Fermer l'onglet 3) Rouvrir l'URL",
           "—",
           "✅ Session restaurée automatiquement, footer affiche le login",
           "Idem"),
        _t("AUTH-06", "—", "Rafraîchissement dynamique post-login",
           "Être sur une page qui change selon le rôle (ex. Caméra, Système)",
           "1) Aller en Guest sur /system 2) Ouvrir le login modal 3) Se connecter en ADMIN",
           "🔁 Avant login : panneau 'État du système' sans disk usage ni menu d'actions",
           "✅ Après login : —",
           "✅ Après login : disk usage + menu reboot/shutdown apparaissent SANS rafraîchir la page"),
    ]),

    ("Page Dashboard (/dashboard)", [
        _t("DASH-01", "/dashboard", "Affichage du tableau de bord par défaut",
           "—",
           "Aller sur /dashboard",
           "✅ Cartes lampe/ventilateur/musique en lecture, capteurs visibles",
           "✅ + boutons d'action fonctionnels",
           "✅ idem USER"),
        _t("DASH-02", "/dashboard", "Switch lampe ON",
           "—",
           "1) Cliquer le toggle 'Lampe' 2) Observer la réponse",
           "⛔ Login modal proposée (HTTP 401 intercepté)",
           "✅ HTTP 200, lampe physique commute, statut mis à jour",
           "✅ idem USER"),
        _t("DASH-03", "/dashboard", "Switch ventilateur ON/OFF",
           "—",
           "Cliquer le toggle 'Ventilateur'",
           "⛔ Login modal",
           "✅ Commute, statut OK",
           "✅"),
        _t("DASH-04", "/dashboard", "Lancer la musique (cocorico ou playlist)",
           "—",
           "Cliquer 'Jouer'",
           "⛔ Login modal",
           "✅",
           "✅"),
        _t("DASH-05", "/dashboard", "Ouverture / fermeture porte",
           "—",
           "Cliquer 'Ouvrir' / 'Fermer'",
           "⛔ Login modal",
           "✅ Commande envoyée (vérifier sur Pi)",
           "✅"),
    ]),

    ("Page Webcam (/camera)", [
        _t("CAM-01", "/camera", "Accès à la page",
           "—",
           "Cliquer 'Webcam' dans le menu OU saisir /camera dans l'URL",
           "⛔ Entrée masquée dans le menu — accès direct URL : redirection vers login",
           "✅ Page chargée",
           "✅ + panneaux admin (réglages caméra, URL d'inférence IA)"),
        _t("CAM-02", "/camera", "Voir le snapshot en direct",
           "Authentifié",
           "Attendre le chargement du snapshot",
           "—",
           "✅ Image affichée (haute qualité ~600 ko)",
           "✅"),
        _t("CAM-03", "/camera", "Analyse IA du snapshot",
           "Authentifié",
           "Cliquer 'Analyse IA'",
           "—",
           "Si ai.inference.url configuré : ✅ message d'analyse. Sinon : message localisé '501 WIP'.",
           "Idem"),
        _t("CAM-04", "/camera", "Navigation dans l'historique des photos",
           "Authentifié",
           "Cliquer un sous-dossier (date) puis une miniature",
           "—",
           "✅ Tri alphabétique ascendant, image agrandie dans une carte dédiée",
           "✅"),
        _t("CAM-05", "/camera", "Modifier la luminosité / rotation",
           "ADMIN",
           "1) Régler le slider luminosité 2) Cliquer 'Save'",
           "—",
           "Panneau de réglages caméra non visible",
           "✅ Toast succès, valeur persistée (vérifier au prochain reboot)"),
        _t("CAM-06", "/camera", "Configurer l'URL d'inférence IA",
           "ADMIN",
           "1) Saisir une URL dans 'AI inference endpoint' 2) Save",
           "—",
           "Panneau non visible",
           "✅ Toast succès, GET /config retourne la nouvelle valeur"),
    ]),

    ("Page Music (/music)", [
        _t("MUS-01", "/music", "Affichage de la page",
           "—",
           "Aller sur /music",
           "✅ Liste des playlists visible",
           "✅",
           "✅"),
        _t("MUS-02", "/music", "Changer la playlist sélectionnée",
           "—",
           "Sélectionner une playlist dans le dropdown puis Save",
           "⛔ 401",
           "✅",
           "✅"),
        _t("MUS-03", "/music", "Configurer 'Sons automatiques' (cocorico, song at sunset)",
           "—",
           "Toggle les deux options puis 'Save'",
           "⛔",
           "✅ Bouton Save apparaît quand dirty + sauvegarde OK",
           "✅"),
        _t("MUS-04", "/music", "Régler le volume",
           "—",
           "Changer le slider de volume puis Save",
           "⛔",
           "✅",
           "✅"),
        _t("MUS-05", "/music", "Layout deux colonnes",
           "—",
           "Sur un écran ≥ lg, vérifier que 'Sons automatiques' et 'Volume' sont côte à côte",
           "✅",
           "✅",
           "✅"),
    ]),

    ("Page Weather (/weather)", [
        _t("WTH-01", "/weather", "Affichage des prévisions",
           "—",
           "Aller sur /weather",
           "✅ Carte météo affichée si configurée",
           "✅",
           "✅"),
        _t("WTH-02", "/weather", "Configurer URL + API key OpenWeather",
           "ADMIN",
           "1) Renseigner URL et key 2) Save",
           "⛔",
           "Panneau visible en lecture (à confirmer) mais sauvegarde 401",
           "✅ Toast succès, key_set passe à true"),
        _t("WTH-03", "/weather", "Mettre à jour latitude / longitude (write-only)",
           "ADMIN",
           "1) Saisir latitude OU longitude 2) Save",
           "—",
           "—",
           "✅ Champs réinitialisés après save. GET /config ne renvoie PAS lat/long (sensible)."),
        _t("WTH-04", "/weather", "Choisir le fournisseur météo",
           "ADMIN",
           "Sélectionner un fournisseur dans le panneau dédié",
           "—",
           "—",
           "✅ Sauvegardé, effet immédiat (hot-reload via cache)"),
    ]),

    ("Page Residents (/residents)", [
        _t("RES-01", "/residents", "Voir la liste des résidentes",
           "—",
           "Aller sur /residents",
           "✅ Liste visible (lecture publique)",
           "✅",
           "✅"),
        _t("RES-02", "/residents", "Ajouter une nouvelle résidente avec photo",
           "USER/ADMIN",
           "1) Bouton 'Ajouter' 2) Remplir 3) Joindre une photo > 1 Mo 4) Save",
           "⛔ Login modal",
           "✅ Photo redimensionnée client-side ~300 ko, upload OK, événement RESIDENT_CREATED",
           "✅"),
        _t("RES-03", "/residents", "Supprimer une résidente",
           "USER/ADMIN",
           "1) Cliquer 'Supprimer' sur une ligne 2) Confirmer dans le modal",
           "⛔",
           "✅ Modal au-dessus de la top-nav (z-index OK), suppression effective, événement RESIDENT_DELETED",
           "✅"),
        _t("RES-04", "/residents", "Upload photo > taille max",
           "USER/ADMIN",
           "Tenter d'uploader une image > 5 Mo brute",
           "—",
           "✅ Compressée avant envoi, pas d'erreur 413",
           "✅"),
    ]),

    ("Page Energy (/energy)", [
        _t("ENG-01", "/energy", "Accès à la page",
           "—",
           "Cliquer 'Energy saving' ou taper /energy",
           "⛔ Menu masqué, URL directe redirige (AdminGuard)",
           "⛔ Idem Guest",
           "✅ Page chargée"),
        _t("ENG-02", "/energy", "Voir le mode de consommation courant",
           "ADMIN",
           "Observer la carte 'Mode courant'",
           "—",
           "—",
           "✅ Mode ECO/REGULAR/SUNNY affiché selon le mapping mensuel"),
        _t("ENG-03", "/energy", "Modifier le mapping mois → mode",
           "ADMIN",
           "1) Changer un mois 2) Save",
           "—",
           "—",
           "✅ Persisté, le mode courant change si on est dans le mois modifié"),
        _t("ENG-04", "/energy", "Forcer le mode ECO",
           "ADMIN",
           "Toggle 'Forcer ECO' puis Save",
           "—",
           "—",
           "✅ Vue 'mode courant' = ECO immédiatement"),
        _t("ENG-05", "/energy", "Contrôles WiFi",
           "ADMIN",
           "Si exposés : tester wifi switch enabled / stop until next door event",
           "—",
           "—",
           "✅ HTTP 200, action loggée. NE PAS exécuter 'stopUntilNextDoorEvent' si on dépend du WiFi pour debug !"),
    ]),

    ("Page Notification / Users (/notification)", [
        _t("USR-01", "/notification", "Accès à la page",
           "—",
           "Aller sur /notification",
           "⛔ AuthGuard redirige",
           "✅ Page chargée",
           "✅"),
        _t("USR-02", "/notification", "Liste des utilisateurs",
           "ADMIN",
           "Observer la liste",
           "—",
           "Liste vide ou erreur 403 sur les endpoints admin",
           "✅ Tous les utilisateurs visibles avec leur rôle"),
        _t("USR-03", "/notification", "Valider un compte NOT_VALIDATED_YET",
           "ADMIN",
           "1) Sélectionner un compte en attente 2) Promouvoir en USER",
           "—",
           "—",
           "✅ Rôle changé, compte peut se connecter"),
        _t("USR-04", "/notification", "Supprimer un compte",
           "ADMIN",
           "1) Clic 'Supprimer' 2) Confirmer",
           "—",
           "—",
           "✅ Compte supprimé"),
        _t("USR-05", "/notification", "S'abonner aux push notifications (Web Push)",
           "USER/ADMIN",
           "Cliquer 'Activer les notifications' (si bouton présent)",
           "—",
           "✅ Permission navigateur demandée, abonnement enregistré",
           "✅ + possibilité de tester via /api/v1/push/test"),
    ]),

    ("Page Electronique (/electronics)", [
        _t("ELE-01", "/electronics", "Voir la table 'Câblage GPIO (BCM)'",
           "—",
           "Aller sur /electronics",
           "✅ 5 colonnes : GPIO / Broche (carte) / Composant / Direction / État",
           "✅ idem + colonne État affiche 'Authentification requise' pour boutons et relais",
           "✅ + colonne État affiche valeurs en direct"),
        _t("ELE-02", "/electronics", "Vérifier le nom des composants",
           "—",
           "Sur la table, vérifier que le label est en FR (compte FR) ou EN (compte EN)",
           "✅",
           "✅",
           "✅"),
        _t("ELE-03", "/electronics", "État en direct des boutons (WebSocket)",
           "ADMIN",
           "1) Appuyer physiquement sur un bouton 2) Observer la cellule 'État' de la ligne correspondante",
           "—",
           "—",
           "✅ Badge passe de 'Released' à 'Pressed' en quasi temps réel + timestamp"),
        _t("ELE-04", "/electronics", "État en direct des relais (HTTP polling)",
           "ADMIN",
           "1) Allumer la lampe depuis /dashboard 2) Aller sur /electronics 3) Attendre ≤ 15 s",
           "—",
           "—",
           "✅ Cellule État de 'Light relay' = badge 'On'"),
        _t("ELE-05", "/electronics", "Image board Raspberry Pi Zero",
           "—",
           "Faire défiler la page",
           "✅ Image visible (chargement externe othermod.com — vérifier qu'elle s'affiche)",
           "✅",
           "✅"),
        _t("ELE-06", "/electronics", "Calibration servomoteur",
           "ADMIN",
           "1) Modifier position ouverte/fermée ou durée 2) Save",
           "—",
           "Panneau visible mais USER+ADMIN seulement",
           "✅ Toast succès, valeur persistée, effet hot-reload"),
        _t("ELE-07", "/electronics", "Nudge clockwise / counter-clockwise",
           "ADMIN",
           "1) Régler durée nudge 2) Clic Counter-clockwise puis Clockwise",
           "—",
           "—",
           "✅ Servo bouge, toasts OK"),
    ]),

    ("Page Logs (/logs)", [
        _t("LOG-01", "/logs", "Voir le journal métier",
           "—",
           "Aller sur /logs, panneau 'Événements métier'",
           "✅ Visible avec libellés localisés (ex. 'Application démarrée', 'Porte fermée')",
           "✅",
           "✅"),
        _t("LOG-02", "/logs", "Filtrer par catégorie et fenêtre temporelle",
           "—",
           "1) Choisir 'Porte' dans le filtre catégorie 2) Choisir '24h'",
           "✅ Filtrage instantané (client-side)",
           "✅",
           "✅"),
        _t("LOG-03", "/logs", "Voir le journal d'authentification",
           "—",
           "Observer le panneau 'Authentifications'",
           "⛔ Panneau masqué (admin only — /events/auth/**)",
           "⛔",
           "✅ Visible avec login success/failed"),
        _t("LOG-04", "/logs", "Voir les fichiers de log",
           "—",
           "Observer le panneau 'Fichiers de log'",
           "⛔ Panneau masqué",
           "⛔",
           "✅ Sélection d'un fichier → contenu affiché avec coloration par niveau"),
        _t("LOG-05", "/logs", "Recherche / filtre dans le fichier de log",
           "ADMIN",
           "1) Sélectionner spring.log 2) Filtrer niveau ERROR 3) Rechercher un mot",
           "—",
           "—",
           "✅ Lignes filtrées affichées"),
    ]),

    ("Page Système (/system)", [
        _t("SYS-01", "/system", "Panneau 'État du système' (anciennement Diagnostic)",
           "—",
           "Aller sur /system",
           "✅ Visible avec version backend (loading puis valeur)",
           "✅ idem",
           "✅ + disk usage + menu actions (... en haut à droite)"),
        _t("SYS-02", "/system", "Panneau 'À propos' en bas de page",
           "—",
           "Faire défiler jusqu'en bas",
           "✅ Affiche auteur, source code, lien Swagger",
           "✅",
           "✅"),
        _t("SYS-03", "/system", "Lien Swagger UI",
           "—",
           "Cliquer 'API documentation (Swagger)' dans le panneau À propos",
           "✅ Ouvre /swagger-ui dans un nouvel onglet (lecture libre)",
           "✅",
           "✅"),
        _t("SYS-04", "/system", "Recharger la configuration (cache evict)",
           "ADMIN",
           "1) Menu ... du panneau État 2) 'Recharger la configuration'",
           "—",
           "—",
           "✅ Toast succès indiquant le nombre de caches vidés"),
        _t("SYS-05", "/system", "Reboot machine",
           "ADMIN",
           "1) Menu ... 2) 'Reboot the machine' — ATTENTION action réelle sur le Pi",
           "—",
           "Entrée masquée du menu",
           "✅ HTTP 200, reboot effectif (à tester avec parcimonie). Rate-limit : 2 appels / 5 min."),
        _t("SYS-06", "/system", "Shutdown machine",
           "ADMIN",
           "Menu ... 'Shut down the machine'",
           "—",
           "—",
           "✅ HTTP 200, audit log enregistré"),
    ]),

    ("i18n / Langue", [
        _t("I18N-01", "—", "Application en français",
           "Locale browser = fr-FR (build localisé)",
           "Charger l'app, parcourir les écrans",
           "✅ Tous les libellés visibles sont traduits",
           "✅",
           "✅"),
        _t("I18N-02", "/logs", "Libellé des événements traduit",
           "—",
           "Vérifier les libellés dans le journal métier",
           "✅ 'Application démarrée', 'Porte ouverte', etc. (pas l'enum brut)",
           "✅",
           "✅"),
        _t("I18N-03", "/electronics", "Nom des composants traduit",
           "—",
           "Vérifier les libellés dans la table GPIO",
           "✅ 'Servomoteur de la porte', 'Bouton de fin de course haut de la porte', etc.",
           "✅",
           "✅"),
    ]),

    ("Sécurité (vérifications transverses)", [
        _t("SEC-01", "DevTools Network", "Vérifier que GET /api/v1/config ne renvoie PAS lat/long",
           "—",
           "1) Se connecter en ADMIN 2) Ouvrir DevTools Network 3) Aller sur /weather 4) Inspecter la réponse JSON",
           "—",
           "—",
           "✅ weather_settings ne contient pas 'latitude' ni 'longitude'"),
        _t("SEC-02", "DevTools Network", "GET /api/v1/config retourne aussi le password seulement en flag",
           "ADMIN",
           "Inspecter email_smtp dans la réponse",
           "—",
           "—",
           "✅ password_set boolean, jamais la valeur"),
        _t("SEC-03", "—", "Rate limit reboot/shutdown",
           "ADMIN",
           "Appeler POST /api/v1/system/reboot 3 fois en < 5 minutes",
           "—",
           "—",
           "✅ 3ème appel : HTTP 429"),
        _t("SEC-04", "—", "CSRF token requis sur POST",
           "USER/ADMIN",
           "1) Se connecter 2) Vérifier que le cookie XSRF-TOKEN est posé "
           "3) Vérifier qu'Angular ajoute X-XSRF-TOKEN aux requêtes mutantes",
           "—",
           "✅",
           "✅"),
        _t("SEC-05", "—", "Audit log sur opérations sensibles",
           "ADMIN",
           "1) Ouvrir/fermer la porte 2) Reboot 3) Consulter le journal d'auth & le logger AUDIT",
           "—",
           "—",
           "✅ Entrées présentes avec login + IP + résultat"),
    ]),
]


def build_test_cases(wb: Workbook) -> None:
    ws = wb.create_sheet("Test cases")

    for i, h in enumerate(TEST_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(TEST_HEADERS))

    row = 2
    for section, tests in TEST_DATA:
        style_section_row(ws, row, len(TEST_HEADERS), section)
        row += 1
        for t in tests:
            for j, v in enumerate(t, start=1):
                ws.cell(row=row, column=j, value=v)
            style_data_row(ws, row, len(TEST_HEADERS))
            ws.row_dimensions[row].height = 65
            row += 1

    # Column widths: Section, ID, URL, scenario, prereq, steps, ex.guest, ex.user, ex.admin, res*3, comment
    set_col_widths(ws, [22, 10, 22, 38, 22, 42, 28, 28, 32, 18, 18, 18, 32])
    ws.freeze_panes = "D2"

    # Data validation for result columns (OK / KO / N/A / Blocked)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"OK,KO,N/A,Blocked"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Valeur invalide"
    dv.errorTitle = "Statut"
    dv.prompt = "OK / KO / N/A / Blocked"
    dv.promptTitle = "Statut"
    ws.add_data_validation(dv)
    # Cover the result columns (J, K, L) for the full range — we use a generous upper bound
    dv.add(f"J2:L{row}")


# ─── Sheet: Smoke ─────────────────────────────────────────────────────────────

def build_smoke(wb: Workbook) -> None:
    ws = wb.create_sheet("Smoke")
    headers = ["#", "Rôle", "Action", "Attendu", "Résultat", "Commentaire"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers))

    rows = [
        ("1", "Guest", "Ouvrir / (dashboard)", "Page chargée, capteurs visibles", "", ""),
        ("2", "Guest", "Cliquer toggle lampe", "Login modal proposée (401)", "", ""),
        ("3", "Guest", "Aller sur /electronics", "Table GPIO 5 colonnes visible, 'Authentification requise' dans État", "", ""),
        ("4", "USER", "Se connecter", "Menu Webcam + Users apparaissent sans F5", "", ""),
        ("5", "USER", "Switch lampe ON", "HTTP 200, statut OK", "", ""),
        ("6", "USER", "Aller sur /energy", "Redirection (AdminGuard) — entrée menu absente", "", ""),
        ("7", "USER", "Se déconnecter", "Footer repasse à 'Guest', éléments USER disparaissent", "", ""),
        ("8", "ADMIN", "Se connecter", "Menu Energy apparaît, panneau État du système avec disk usage", "", ""),
        ("9", "ADMIN", "Recharger la config (menu ...)", "Toast 'X caches vidés'", "", ""),
        ("10", "ADMIN", "Vérifier journal d'auth (/logs)", "Panneau Auth events visible, LOGIN_SUCCESS récent listé", "", ""),
    ]
    for i, r in enumerate(rows, start=2):
        for j, v in enumerate(r, start=1):
            ws.cell(row=i, column=j, value=v)
        style_data_row(ws, i, len(headers))
        ws.row_dimensions[i].height = 32

    set_col_widths(ws, [5, 10, 35, 50, 14, 35])
    ws.freeze_panes = "A2"

    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"OK,KO,N/A,Blocked"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"E2:E{len(rows) + 1}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = Workbook()
    # Remove the default sheet — we create our own.
    wb.remove(wb.active)

    build_cover(wb)
    build_roles(wb)
    build_test_cases(wb)
    build_smoke(wb)

    # Reorder: Cover first
    order = ["Cover", "Roles", "Test cases", "Smoke"]
    wb._sheets = [wb[name] for name in order]

    out = "docs/test-plan/hermanas-test-plan.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
